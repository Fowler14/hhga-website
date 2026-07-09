#!/usr/bin/env python3
"""HHGA data extraction — reads the HH year folders (READ ONLY) and writes data/hhga.json.

Each trip year has its own parser because the early workbooks all differ.
Run:  python3 scripts/extract.py
"""

import json
import re
from pathlib import Path

import xlrd

ROOT = Path(__file__).resolve().parent.parent
HH = ROOT / "HH"
YEARS_DIR = ROOT / "data" / "years"
OUT = ROOT / "data" / "hhga.json"
ALIASES = ROOT / "data" / "aliases.json"


def norm_name(name):
    """Normalize a workbook name for alias lookup: lowercase, collapse
    whitespace, no space after commas ('Fowler,P ' == 'fowler,p')."""
    return re.sub(r",\s*", ",", re.sub(r"\s+", " ", name.strip().lower()))


def load_aliases():
    """alias (as written in a workbook) -> canonical player id."""
    with open(ALIASES) as f:
        players = json.load(f)["players"]
    lookup = {}
    for p in players:
        for alias in p["aliases"] + [p["name"]]:
            lookup[norm_name(alias)] = p["id"]
    return players, lookup


def xl_date(wb, value):
    """Excel serial date -> ISO string."""
    y, m, d, *_ = xlrd.xldate.xldate_as_tuple(value, wb.datemode)
    return f"{y:04d}-{m:02d}-{d:02d}"


def to_int(v):
    return int(v) if isinstance(v, float) else None


# ---------------------------------------------------------------- 2000

def extract_2000(alias_lookup):
    """2000: 'HH golf scores.xls', sheet 'Scores WPar'.

    5 rounds x 27 holes, 8 players. Gross scores only — no handicap/net
    data exists for this year. Round 1 also has 4some groupings and
    Skins/Greenies/Sandies tallies.
    Layout: one row per (round, player); PAR appears as a pseudo-player row.
    Cols: 0 Round, 1 4Some, 2 Course, 3 Player, 4 Date (excel serial),
          5-13 holes 1-9, 14 out-total, 15-23 holes 10-18, 24 in-total,
          25 18-tot, 26-34 holes 19-27, 35 third-nine-total, 36 27-tot,
          37 Skins, 38 Greenies, 39 Sandies.
    """
    wb = xlrd.open_workbook(HH / "2000" / "HH golf scores.xls")
    s = wb.sheet_by_name("Scores WPar")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24)) + list(range(26, 35))

    rounds = {}  # round number -> round dict
    side = {}    # player id -> {skins, greenies, sandies} (round 1 only)

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        rnd = to_int(s.cell_value(r, 0))
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        stated_total = to_int(s.cell_value(r, 36))
        assert sum(holes) == stated_total, (
            f"2000 R{rnd} {name}: holes sum {sum(holes)} != stated {stated_total}")

        if rnd not in rounds:
            rounds[rnd] = {
                "round": rnd,
                "course": s.cell_value(r, 2),
                "date": xl_date(wb, s.cell_value(r, 4)),
                "holes": 27,
                "par": None,
                "scores": {},
                "foursomes": {},
            }

        if name == "PAR":
            rounds[rnd]["par"] = holes
        else:
            pid = alias_lookup[norm_name(name)]
            rounds[rnd]["scores"][pid] = holes
            foursome = to_int(s.cell_value(r, 1))
            if foursome:
                rounds[rnd]["foursomes"][pid] = foursome
            if rnd == 1:
                side[pid] = {
                    "skins": to_int(s.cell_value(r, 37)),
                    "greenies": to_int(s.cell_value(r, 38)),
                    "sandies": to_int(s.cell_value(r, 39)),
                }

    rounds = [rounds[k] for k in sorted(rounds)]
    for rd in rounds:
        if not rd["foursomes"]:
            del rd["foursomes"]

    # Gross standings across all 5 rounds (no net data exists in 2000).
    totals = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            totals[pid] = totals.get(pid, 0) + sum(holes)
    board = sorted(totals.items(), key=lambda kv: kv[1])
    leaderboard = [
        {"player": pid, "gross": tot, "place": i + 1}
        for i, (pid, tot) in enumerate(board)
    ]

    return {
        "year": 2000,
        "location": "Hilton Head Island, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "gross",  # no handicap/net data this year
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": "Verified by Ryan 2026-07-02. No handicap/net data exists for 2000; champion = lowest gross.",
        "sideGames": {"round1": side},
    }


# ---------------------------------------------------------------- 2001

def extract_2001(alias_lookup):
    """2001: 'HH golf scores 2001.xls', sheet 'Scores WPar'.

    6 rounds x 27 holes, 8 players (John Taber left after round 3).
    First year with handicaps: cols 37-39 are Handicap / 18 net / 27 net
    (verified: 27net = 27gross - 1.5*handicap on every row). Handicaps were
    derived from the 2000 scores (sheet '2000 Official Hndcps').
    No skins/greenies/sandies this year.
    Dates: round 1 serial says 2001-02-18 but rounds 2-6 say 2000-02-19..23 —
    a template-copy typo; the trip ran consecutive days, so we correct the
    year to 2001 (Feb 18-23).
    PAR rows for rounds 2-6 have a blank course cell; course is taken from
    the player rows.
    """
    wb = xlrd.open_workbook(HH / "2001" / "HH golf scores 2001.xls")
    s = wb.sheet_by_name("Scores WPar")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24)) + list(range(26, 35))

    rounds = {}
    handicaps = {}
    par_notes = []

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        rnd = to_int(s.cell_value(r, 0))
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        stated_total = to_int(s.cell_value(r, 36))
        if name == "PAR":
            # R3 (Hilton Head National) PAR: holes + nine-subtotals all agree
            # on 36+36+35=107, but the 27-hole total cell says 108. The hole
            # detail wins; the workbook's grand-total cell is a typo.
            if sum(holes) != stated_total:
                par_notes.append(
                    f"Round {rnd}: PAR total cell says {stated_total}, "
                    f"hole-by-hole par sums to {sum(holes)} (used the holes).")
        else:
            assert sum(holes) == stated_total, (
                f"2001 R{rnd} {name}: holes sum {sum(holes)} != stated {stated_total}")

        if rnd not in rounds:
            # Workbook date bug: rounds 2-6 kept year 2000 from the copied
            # 2000 template. Trip = consecutive days from 2001-02-18.
            rounds[rnd] = {
                "round": rnd,
                "course": None,
                "date": f"2001-02-{17 + rnd:02d}",
                "holes": 27,
                "par": None,
                "scores": {},
                "foursomes": {},
            }

        course = s.cell_value(r, 2)
        if course and not rounds[rnd]["course"]:
            rounds[rnd]["course"] = course

        if name == "PAR":
            rounds[rnd]["par"] = holes
        else:
            pid = alias_lookup[norm_name(name)]
            rounds[rnd]["scores"][pid] = holes
            handicaps[pid] = to_int(s.cell_value(r, 37))
            net = s.cell_value(r, 39)
            assert abs(stated_total - 1.5 * handicaps[pid] - net) < 1e-9, (
                f"2001 R{rnd} {name}: net {net} != {stated_total} - 1.5*{handicaps[pid]}")
            foursome = to_int(s.cell_value(r, 1))
            if foursome:
                rounds[rnd]["foursomes"][pid] = foursome

    rounds = [rounds[k] for k in sorted(rounds)]
    for rd in rounds:
        if not rd["foursomes"]:
            del rd["foursomes"]

    # Standings: net total = gross - 1.5*handicap per 27-hole round.
    gross, played = {}, {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)
            played[pid] = played.get(pid, 0) + 1
    n_rounds = len(rounds)
    complete = [(pid, g, g - 1.5 * handicaps[pid] * played[pid])
                for pid, g in gross.items() if played[pid] == n_rounds]
    complete.sort(key=lambda t: t[2])
    leaderboard = [
        {"player": pid, "gross": g, "handicap": handicaps[pid],
         "net": net, "place": i + 1}
        for i, (pid, g, net) in enumerate(complete)
    ]
    for pid, g in gross.items():  # incomplete trips go at the bottom, no place
        if played[pid] < n_rounds:
            leaderboard.append({
                "player": pid, "gross": g, "handicap": handicaps[pid],
                "net": g - 1.5 * handicaps[pid] * played[pid], "place": None,
                "roundsPlayed": played[pid],
            })

    return {
        "year": 2001,
        "location": "Hilton Head Island, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",  # first year with handicaps (from 2000 scores)
        "rounds": rounds,
        "handicaps": handicaps,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-02: champion = lowest net "
                         "(Tom Conroy 832), not the gross ranking in the "
                         "workbook's '2001 Total Strokes' sheet."),
        "dateNote": "Rounds 2-6 dated year 2000 in the workbook (template typo); corrected to 2001.",
        "parNotes": par_notes,
    }


# ---------------------------------------------------------------- 2002

def extract_2002(alias_lookup):
    """2002: 'HH golf scores 2002.xls', sheet '2002 scores'.

    5 rounds x 27 holes, Mar 12-16 2002, 8 players: Jose Perna out,
    Phil Fowler's first year. Same column layout as 2001 (Handicap /
    18 net / 27 net). Handicaps computed from each player's 2000+2001
    rounds (sheet '2002 hndcps': average-per-18 dropping the high, x1.5).
    Quirks:
      - Rounds are numbered 1,2,3,4,6 in the sheet despite consecutive
        dates — renumbered here to 1-5.
      - Phil Fowler (new) has handicap 0.
      - Dan Taber's handicap is 14 in rounds 1-4 but 0 in the final round;
        the workbook's net column applies the 0. Kept as the workbook has
        it (flagged in handicapNotes).
    Separate winnings file: 'HH 2002 Tournament Winnings.xls' (daily
    money per player, Phil Fowler +116.25 on top).
    """
    wb = xlrd.open_workbook(HH / "2002" / "HH golf scores 2002.xls")
    s = wb.sheet_by_name("2002 scores")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24)) + list(range(26, 35))

    rounds = {}
    handicaps = {}      # pid -> most common handicap for the year
    round_nets = {}     # pid -> summed 27-hole net as the workbook applied it
    hcp_notes = []

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        rnd = to_int(s.cell_value(r, 0))
        if rnd == 6:  # sheet numbers the 5th day "6"; dates are consecutive
            rnd = 5
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        stated_total = to_int(s.cell_value(r, 36))
        assert sum(holes) == stated_total, (
            f"2002 R{rnd} {name}: holes sum {sum(holes)} != stated {stated_total}")

        if rnd not in rounds:
            rounds[rnd] = {
                "round": rnd,
                "course": None,
                "date": xl_date(wb, s.cell_value(r, 4)),
                "holes": 27,
                "par": None,
                "scores": {},
                "foursomes": {},
            }
        course = s.cell_value(r, 2)
        if course and not rounds[rnd]["course"]:
            rounds[rnd]["course"] = course

        if name == "PAR":
            rounds[rnd]["par"] = holes
        else:
            pid = alias_lookup[norm_name(name)]
            rounds[rnd]["scores"][pid] = holes
            hcp = to_int(s.cell_value(r, 37))
            net = s.cell_value(r, 39)
            assert abs(stated_total - 1.5 * hcp - net) < 1e-9, (
                f"2002 R{rnd} {name}: net {net} != {stated_total} - 1.5*{hcp}")
            round_nets[pid] = round_nets.get(pid, 0) + net
            if pid in handicaps and handicaps[pid] != hcp:
                hcp_notes.append(
                    f"{pid}: handicap {handicaps[pid]} in earlier rounds but "
                    f"{hcp} in round {rnd} — workbook applied as-is.")
            else:
                handicaps[pid] = hcp
            foursome = to_int(s.cell_value(r, 1))
            if foursome:
                rounds[rnd]["foursomes"][pid] = foursome

    rounds = [rounds[k] for k in sorted(rounds)]
    for rd in rounds:
        if not rd["foursomes"]:
            del rd["foursomes"]

    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)

    board = sorted(round_nets.items(), key=lambda kv: kv[1])
    leaderboard = [
        {"player": pid, "gross": gross[pid], "handicap": handicaps[pid],
         "net": net, "place": i + 1}
        for i, (pid, net) in enumerate(board)
    ]

    # Daily money from the separate winnings workbook.
    wwb = xlrd.open_workbook(HH / "2002" / "HH 2002 Tournament Winnings.xls")
    ws = wwb.sheet_by_name("Sheet1")
    winnings = {}
    for r in range(1, ws.nrows):
        name = ws.cell_value(r, 1)
        if name and name != "Variance":
            winnings[alias_lookup[norm_name(name)]] = ws.cell_value(r, 7)

    return {
        "year": 2002,
        "location": "Hilton Head Island, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "rounds": rounds,
        "handicaps": handicaps,
        "handicapNotes": hcp_notes,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-02: champion = lowest net "
                         "(Joe Dueh 578.5). Phil Fowler had the lowest gross (648) "
                         "but played off 0 as a first-year player."),
        "roundNote": "Sheet numbers the final round '6' (no round 5); renumbered to 5 — dates are consecutive Mar 12-16.",
        "sideGames": {"winnings": winnings},
    }


# ---------------------------------------------------------------- 2003

def extract_2003(alias_lookup):
    """2003: 'HH Score 2003.xls', sheet '2003 Scores'.

    5 rounds, Feb 25 - Mar 1 2003, 7 players (Joe Dueh absent).
    Rounds 1-4 are 27 holes; round 5 (Arthur Hill) is 18 holes — the
    holes-19-27 cells are empty. The workbook still applies the 27-hole
    net formula (gross - 1.5*handicap) to the 18-hole round; the year
    total ('2003 Total' sheet) sums those nets, so we keep that math.
    Course-name typos kept as written ('Hilton Head Naational'), only
    whitespace collapsed.
    """
    wb = xlrd.open_workbook(HH / "2003" / "HH Score 2003.xls")
    s = wb.sheet_by_name("2003 Scores")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24)) + list(range(26, 35))

    rounds = {}
    handicaps = {}
    round_nets = {}

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        rnd = to_int(s.cell_value(r, 0))
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        # 18-hole rounds leave holes 19-27 as 0s (blank cells read as 0.0)
        if all(not h for h in holes[18:]):
            holes = holes[:18]
        assert all(h for h in holes), f"2003 R{rnd} {name}: hole gaps"
        stated_total = to_int(s.cell_value(r, 36))
        assert sum(holes) == stated_total, (
            f"2003 R{rnd} {name}: holes sum {sum(holes)} != stated {stated_total}")

        if rnd not in rounds:
            rounds[rnd] = {
                "round": rnd,
                "course": None,
                "date": xl_date(wb, s.cell_value(r, 4)),
                "holes": len(holes),
                "par": None,
                "scores": {},
                "foursomes": {},
            }
        course = s.cell_value(r, 2)
        if course and not rounds[rnd]["course"]:
            rounds[rnd]["course"] = re.sub(r"\s+", " ", course.strip())

        if name == "PAR":
            rounds[rnd]["par"] = holes
        else:
            pid = alias_lookup[norm_name(name)]
            rounds[rnd]["scores"][pid] = holes
            hcp = to_int(s.cell_value(r, 37))
            net = s.cell_value(r, 39)
            assert abs(stated_total - 1.5 * hcp - net) < 1e-9, (
                f"2003 R{rnd} {name}: net {net} != {stated_total} - 1.5*{hcp}")
            handicaps[pid] = hcp
            round_nets[pid] = round_nets.get(pid, 0) + net
            foursome = to_int(s.cell_value(r, 1))
            if foursome:
                rounds[rnd]["foursomes"][pid] = foursome

    rounds = [rounds[k] for k in sorted(rounds)]
    for rd in rounds:
        if not rd["foursomes"]:
            del rd["foursomes"]

    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)

    board = sorted(round_nets.items(), key=lambda kv: kv[1])
    leaderboard = [
        {"player": pid, "gross": gross[pid], "handicap": handicaps[pid],
         "net": net, "place": i + 1}
        for i, (pid, net) in enumerate(board)
    ]

    return {
        "year": 2003,
        "location": "Hilton Head Island, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "rounds": rounds,
        "handicaps": handicaps,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-02: champion = lowest net "
                         "(John Williams 544.5, hcp 25)."),
        "formatNote": ("Rounds 1-4 are 27 holes; round 5 (Arthur Hill) is 18. The "
                       "workbook applies net = gross - 1.5*handicap to every round "
                       "including the 18-hole one; year total follows the workbook."),
    }


# ---------------------------------------------------------------- 2004

def extract_2004(alias_lookup):
    """2004: 'HH Score 2004 new format.xls', sheet '2004_Scores'.

    First Myrtle Beach year. 6 rounds, Mar 9-13 2004, 8 players.
    Names switch to 'Last, First-initial' ('Conroy, T', 'Fowler,P').
    New: Dan Travisano. 'Dueh, Jr.' = Joe Dueh (Ryan confirmed same
    person). John Taber and Jose Perna absent.

    Round 1 = 27 holes split over King's North + King's South; rounds
    2-6 = 18 holes. Two rounds on Mar 11 (River Club + Willbrook).
    The sheet also contains zero-filled 'MBN South' blocks for planned-
    but-unplayed holes on Mar 9 and Mar 13 — skipped (first 18 holes
    all zero). Net = gross - 1.5*handicap on every round (verified);
    the year total ('2004_Total', '27 Hole Net' table) sums those.

    A second file 'HH Score 2004 Workcopy.xls' differs only in round
    numbering (labels Willbrook round 3); scores are identical.
    """
    wb = xlrd.open_workbook(HH / "2004" / "HH Score 2004 new format.xls")
    s = wb.sheet_by_name("2004_Scores")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24)) + list(range(26, 35))

    rounds = {}
    handicaps = {}
    round_nets = {}

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        rnd = to_int(s.cell_value(r, 0))
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        if not any(holes[:18]):  # unplayed 'MBN South' filler block
            continue
        if all(not h for h in holes[18:]):  # 18-hole round
            holes = holes[:18]
        assert all(h for h in holes), f"2004 R{rnd} {name}: hole gaps"
        stated_total = to_int(s.cell_value(r, 36))
        assert sum(holes) == stated_total, (
            f"2004 R{rnd} {name}: holes sum {sum(holes)} != stated {stated_total}")

        if rnd not in rounds:
            rounds[rnd] = {
                "round": rnd,
                "course": None,
                "date": xl_date(wb, s.cell_value(r, 4)),
                "holes": len(holes),
                "par": None,
                "scores": {},
                "foursomes": {},
            }
        course = s.cell_value(r, 2)
        if course and not rounds[rnd]["course"]:
            course = re.sub(r"\s+", " ", course.strip())
            if course == "Lichtfield":  # workbook typo, Ryan confirmed spelling
                course = "Litchfield"
            rounds[rnd]["course"] = course

        if name == "PAR":
            rounds[rnd]["par"] = holes
        else:
            pid = alias_lookup[norm_name(name)]
            rounds[rnd]["scores"][pid] = holes
            hcp = to_int(s.cell_value(r, 37))
            net = s.cell_value(r, 39)
            assert abs(stated_total - 1.5 * hcp - net) < 1e-9, (
                f"2004 R{rnd} {name}: net {net} != {stated_total} - 1.5*{hcp}")
            handicaps[pid] = hcp
            round_nets[pid] = round_nets.get(pid, 0) + net
            foursome = to_int(s.cell_value(r, 1))
            if foursome:
                rounds[rnd]["foursomes"][pid] = foursome

    rounds = [rounds[k] for k in sorted(rounds)]
    for rd in rounds:
        if not rd["foursomes"]:
            del rd["foursomes"]
        # PAR rows carry 27 values even for 18-hole rounds (phantom third
        # nine); size each round by what the players actually played.
        n = max(len(h) for h in rd["scores"].values())
        assert all(len(h) == n for h in rd["scores"].values()), (
            f"2004 R{rd['round']}: players played differing hole counts")
        rd["holes"] = n
        rd["par"] = rd["par"][:n]
    # Round 1 was 27 holes across two courses.
    rounds[0]["course"] = "King's North + King's South"

    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)

    board = sorted(round_nets.items(), key=lambda kv: kv[1])
    leaderboard = [
        {"player": pid, "gross": gross[pid], "handicap": handicaps[pid],
         "net": net, "place": i + 1}
        for i, (pid, net) in enumerate(board)
    ]

    # Winnings file uses full/nick names ('Pappy Williams' = John Williams).
    wwb = xlrd.open_workbook(HH / "2004" / "HH 2004 Tournament Winnings.xls")
    ws = wwb.sheet_by_name("Sheet1")
    winnings = {}
    for r in range(1, ws.nrows):
        name = ws.cell_value(r, 1)
        if name and name != "Variance":
            winnings[alias_lookup[norm_name(name)]] = ws.cell_value(r, 8)

    return {
        "year": 2004,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "rounds": rounds,
        "handicaps": handicaps,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-02: champion = lowest net "
                         "(Jim Fowler 555, one ahead of Phil Fowler 556)."),
        "rosterNote": ("Ryan confirmed: 'Dueh, Jr.' = Joe Dueh (same person); "
                       "'Pappy Williams' = John Williams. 36-hole day Mar 11 "
                       "confirmed real."),
        "sideGames": {"winnings": winnings},
    }


# ---------------------------------------------------------------- 2005

def extract_2005(alias_lookup):
    """2005: 'HH Score 2005 Working.xls', sheet '2005_Scores'.

    Myrtle Beach again. 5 rounds x 18 holes, 12 players — the nickname
    era begins (TC, Gregory, Joey, Jimmy, Phil, Dano, Trav, Pappy,
    Johnny T back, plus new: Duck, Bruce, Big Gay Al).

    Sheet quirks:
      - Blocks are delimited by PAR rows; the first two blocks are BOTH
        labelled round 1 and the extra-players' rows have blank round
        cells — rounds are therefore assigned by PAR-row order.
      - Dates are 2004 template leftovers on most rows; only round 1
        (2005-03-08) is trustworthy. Rounds dated consecutively from
        there, flagged as estimated.
      - Net = gross - handicap per 18-hole round (col 38); the '27 net'
        column still computes gross - 1.5*hcp but is meaningless this
        year and ignored. Verified against the '2005_Total' net table.
      - No tournament winnings file for 2005 (only a one-day Wolf-game
        card calculator).
    """
    wb = xlrd.open_workbook(HH / "2005" / "HH Score 2005 Working.xls")
    s = wb.sheet_by_name("2005_Scores")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24)) + list(range(26, 35))

    rounds = []
    handicaps = {}
    round_nets = {}

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        if all(not h for h in holes[18:]):
            holes = holes[:18]
        assert all(h for h in holes), f"2005 {name}: hole gaps"
        stated_total = to_int(s.cell_value(r, 36))
        assert sum(holes) == stated_total, (
            f"2005 {name}: holes sum {sum(holes)} != stated {stated_total}")

        if name == "PAR":  # every PAR row starts the next round
            rounds.append({
                "round": len(rounds) + 1,
                "course": re.sub(r"\s+", " ", s.cell_value(r, 2).strip()),
                "date": f"2005-03-{7 + len(rounds) + 1:02d}",
                "holes": len(holes),
                "par": holes,
                "scores": {},
                "foursomes": {},
            })
        else:
            rd = rounds[-1]
            pid = alias_lookup[norm_name(name)]
            rd["scores"][pid] = holes
            hcp = to_int(s.cell_value(r, 37))
            net = s.cell_value(r, 38)  # 18-hole net this year
            assert abs(stated_total - hcp - net) < 1e-9, (
                f"2005 {name}: net {net} != {stated_total} - {hcp}")
            handicaps[pid] = hcp
            round_nets[pid] = round_nets.get(pid, 0) + net
            foursome = to_int(s.cell_value(r, 1))
            if foursome:
                rd["foursomes"][pid] = foursome

    for rd in rounds:
        if not rd["foursomes"]:
            del rd["foursomes"]

    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)

    # Tie-aware places (Phil/Pappy/Duck all net 448 this year).
    board = sorted(round_nets.items(), key=lambda kv: kv[1])
    leaderboard = []
    for i, (pid, net) in enumerate(board):
        place = leaderboard[-1]["place"] if leaderboard and leaderboard[-1]["net"] == net else i + 1
        leaderboard.append({"player": pid, "gross": gross[pid],
                            "handicap": handicaps[pid], "net": net,
                            "place": place})

    return {
        "year": 2005,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "rounds": rounds,
        "handicaps": handicaps,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-02: champion = lowest net "
                         "(Greg Dueh 418). Duck = Don Fowler, Bruce = Bruce Morris. "
                         "Big Gay Al was a one-time stand-in (real name unknown), "
                         "not Roy Hoenisch."),
        "dateNote": ("Only round 1's date (2005-03-08) is reliable; later rows kept "
                     "2004 template dates. Rounds dated as consecutive days Mar 8-12 "
                     "— estimated, Ryan to confirm (there may have been a 36-hole day)."),
    }


# ---------------------------------------------------------------- 2006

def extract_2006(alias_lookup):
    """2006: 'HH Score 2006 working.xls', sheet '2006_Scores'.

    Myrtle Beach, 5 rounds x 18, same crew as 2005 except Roy Hoenisch
    ('Big Gay Roy', hcp 0) replaces the 2005 stand-in 'Big Gay Al' —
    different people; Roy's nickname was a joke about his predecessor.

    First year of ROLLING HANDICAPS ('Rolling' sheet): each day's
    handicap = f(best 3 of last 6 rounds), seeded from 2005, so the
    handicap column varies round to round — stored per round here.
    Net = gross - that day's handicap; year total = sum of nets,
    verified against the sheet's own Leaderboard block (first year one
    exists): Trav 426 on top.

    Round dates are template leftovers again (2005/2004); money files
    suggest the trip ran ~Mar 5-9 2006 — dated consecutively, estimated.
    """
    wb = xlrd.open_workbook(HH / "2006" / "HH Score 2006 working.xls")
    s = wb.sheet_by_name("2006_Scores")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24)) + list(range(26, 35))

    rounds = []
    round_nets = {}

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        if all(not h for h in holes[18:]):
            holes = holes[:18]
        assert all(h for h in holes), f"2006 {name}: hole gaps"
        stated_total = to_int(s.cell_value(r, 36))
        assert sum(holes) == stated_total, (
            f"2006 {name}: holes sum {sum(holes)} != stated {stated_total}")

        if name == "PAR":
            rounds.append({
                "round": len(rounds) + 1,
                "course": re.sub(r"\s+", " ", s.cell_value(r, 2).strip()),
                "date": f"2006-03-{4 + len(rounds) + 1:02d}",
                "holes": len(holes),
                "par": holes,
                "scores": {},
                "handicaps": {},  # rolling: varies per round
            })
        else:
            rd = rounds[-1]
            pid = alias_lookup[norm_name(name)]
            hcp = to_int(s.cell_value(r, 37))
            net = s.cell_value(r, 38)
            assert abs(stated_total - hcp - net) < 1e-9, (
                f"2006 {name}: net {net} != {stated_total} - {hcp}")
            rd["scores"][pid] = holes
            rd["handicaps"][pid] = hcp
            round_nets[pid] = round_nets.get(pid, 0) + net

    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)

    board = sorted(round_nets.items(), key=lambda kv: kv[1])
    leaderboard = []
    for i, (pid, net) in enumerate(board):
        place = leaderboard[-1]["place"] if leaderboard and leaderboard[-1]["net"] == net else i + 1
        leaderboard.append({"player": pid, "gross": gross[pid],
                            "net": net, "place": place})

    return {
        "year": 2006,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": ("Rolling daily handicaps (first year): each day's "
                           "handicap recomputed from best 3 of the player's "
                           "last 6 rounds, seeded from 2005. Stored per round."),
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-02: champion = lowest net "
                         "(Dan Travisano 426)."),
        "dateNote": ("Score-sheet dates are template leftovers; trip dated ~Mar 5-9 "
                     "2006 from the money-calculator files — estimated, Ryan to confirm."),
    }


# ---------------------------------------------------------------- 2007

def extract_2007(alias_lookup):
    """2007: 'HH Score 2007 Working.xls', sheet '2007_Scores'.

    Myrtle Beach, 5 rounds x 18, Mar 12-16 2007 (dates real this year).
    Roster: Greg Dueh, Don Fowler and Roy out; NEW: Irv, Craig, and
    Billy ('Brother Billy', played Glen Dornoch only — his other rows
    are gross 0 with formula-artifact negative handicaps and the
    Total sheet's 227 'net' for him is meaningless; he gets no place).
    Rolling handicaps continue but recalibrated much higher
    (Pappy up to 40). Net = gross - day's handicap (col 38).
    """
    wb = xlrd.open_workbook(HH / "2007" / "HH Score 2007 Working.xls")
    s = wb.sheet_by_name("2007_Scores")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24)) + list(range(26, 35))

    rounds = []
    round_nets = {}
    played = {}

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        stated_total = to_int(s.cell_value(r, 36))
        if name != "PAR" and not stated_total:
            continue  # gross 0 = didn't play that day (Billy)
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        if all(not h for h in holes[18:]):
            holes = holes[:18]
        assert all(h for h in holes), f"2007 {name}: hole gaps"
        assert sum(holes) == stated_total, (
            f"2007 {name}: holes sum {sum(holes)} != stated {stated_total}")

        if name == "PAR":
            rounds.append({
                "round": len(rounds) + 1,
                "course": re.sub(r"\s+", " ", s.cell_value(r, 2).strip()),
                "date": xl_date(wb, s.cell_value(r, 4)),
                "holes": len(holes),
                "par": holes,
                "scores": {},
                "handicaps": {},
            })
        else:
            rd = rounds[-1]
            pid = alias_lookup[norm_name(name)]
            hcp = to_int(s.cell_value(r, 37))
            net = s.cell_value(r, 38)
            assert abs(stated_total - hcp - net) < 1e-9, (
                f"2007 {name}: net {net} != {stated_total} - {hcp}")
            rd["scores"][pid] = holes
            rd["handicaps"][pid] = hcp
            round_nets[pid] = round_nets.get(pid, 0) + net
            played[pid] = played.get(pid, 0) + 1

    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)

    n_rounds = len(rounds)
    board = sorted(
        ((pid, net) for pid, net in round_nets.items() if played[pid] == n_rounds),
        key=lambda kv: kv[1])
    leaderboard = []
    for i, (pid, net) in enumerate(board):
        place = leaderboard[-1]["place"] if leaderboard and leaderboard[-1]["net"] == net else i + 1
        leaderboard.append({"player": pid, "gross": gross[pid],
                            "net": net, "place": place})
    for pid, cnt in played.items():
        if cnt < n_rounds:
            leaderboard.append({"player": pid, "gross": gross[pid],
                                "net": round_nets[pid], "place": None,
                                "roundsPlayed": cnt})

    return {
        "year": 2007,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (recalibrated higher than 2006).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-02: champion = lowest net "
                         "(John Taber 359, by one over John Williams). Billy "
                         "(1 round) excluded from places."),
        "rosterNote": ("New in 2007: Irv, Craig, Billy ('Brother Billy', 1 round) — "
                       "full names needed from Ryan."),
    }


# ---------------------------------------------------------------- 2008

def extract_2008(alias_lookup):
    """2008: 'HH Score 2008  Working.xls', sheet '2008_Scores'.

    Myrtle Beach, 5 rounds x 18, Mar 10-14 2008 (only round 1's date is
    real; the rest are 2007 template leftovers — dated consecutively).
    Roster back to the 2006 crew: Gregory, Duck, Big Gay Roy return;
    Irv and Billy out. 12 players.

    New 18-hole-native layout: cols 5-23 holes, 25 gross, 26 handicap
    (rolling, per round), 27 net, 28 pars, 29 birdies (blank = 0).
    First year with a dedicated 'LeaderBoard' sheet (Name/Net/Diff/
    Place) — places taken from it verbatim (it breaks the Dano/Craig
    375 tie by giving them places 2 and 3), nets cross-checked against
    summed per-round nets.
    """
    wb = xlrd.open_workbook(HH / "2008" / "HH Score 2008  Working.xls")
    s = wb.sheet_by_name("2008_Scores")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24))

    rounds = []
    round_nets = {}
    stats = {}  # pid -> {pars, birdies}

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        assert all(h for h in holes), f"2008 {name}: hole gaps"
        stated_total = to_int(s.cell_value(r, 25))
        assert sum(holes) == stated_total, (
            f"2008 {name}: holes sum {sum(holes)} != stated {stated_total}")

        if name == "PAR":
            rounds.append({
                "round": len(rounds) + 1,
                "course": re.sub(r"\s+", " ", s.cell_value(r, 2).strip()),
                "date": f"2008-03-{9 + len(rounds) + 1:02d}",
                "holes": 18,
                "par": holes,
                "scores": {},
                "handicaps": {},
            })
        else:
            rd = rounds[-1]
            pid = alias_lookup[norm_name(name)]
            hcp = to_int(s.cell_value(r, 26))
            net = s.cell_value(r, 27)
            assert abs(stated_total - hcp - net) < 1e-9, (
                f"2008 {name}: net {net} != {stated_total} - {hcp}")
            rd["scores"][pid] = holes
            rd["handicaps"][pid] = hcp
            round_nets[pid] = round_nets.get(pid, 0) + net
            st = stats.setdefault(pid, {"pars": 0, "birdies": 0})
            st["pars"] += to_int(s.cell_value(r, 28)) or 0
            st["birdies"] += to_int(s.cell_value(r, 29)) or 0

    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)

    # Places straight from the workbook's LeaderBoard sheet.
    lb = wb.sheet_by_name("LeaderBoard")
    leaderboard = []
    for r in range(1, lb.nrows):
        name = lb.cell_value(r, 0)
        if not name:
            continue
        pid = alias_lookup[norm_name(name)]
        net = lb.cell_value(r, 1)
        assert abs(round_nets[pid] - net) < 1e-9, (
            f"2008 LeaderBoard {name}: sheet net {net} != summed {round_nets[pid]}")
        leaderboard.append({"player": pid, "gross": gross[pid], "net": net,
                            "place": to_int(lb.cell_value(r, 3)),
                            "pars": stats[pid]["pars"],
                            "birdies": stats[pid]["birdies"]})

    return {
        "year": 2008,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps.",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-02: champion = Joe Dueh (359), "
                         "per the LeaderBoard sheet (which breaks the 375 tie: "
                         "Dano 2nd, Craig 3rd)."),
        "dateNote": ("Only round 1's date (2008-03-10) is real; rounds dated "
                     "consecutively Mar 10-14, estimated."),
        "sideGames": {"myrtleCup": myrtle_cup_2008()},
    }


def myrtle_cup_2008():
    """Transcribed from 'HH/2008/Myrtle Cup 2008.pdf' — the first recorded
    Mist vs Weed team event ('Myrtle Cup'). The doc is a mid-trip snapshot
    ('Results after Wednesday March 12'): 4.5-4.5 with Thursday singles
    still to play; the final result is not in the folder.
    'Knees' = Craig; 'JW' = John Williams."""
    return {
        "name": "Myrtle Cup",
        "format": "6v6 team match play: doubles Mon-Wed, singles Thu; 15 points total, 8 to win",
        "teams": {
            "mist": ["phil-fowler", "don-fowler", "john-williams",
                     "roy-hoenisch", "tom-conroy", "joe-dueh"],
            "weed": ["dan-taber", "greg-dueh", "john-taber",
                     "dan-travisano", "jim-fowler", "craig"],
        },
        "standingsAfterWed": {"mist": 4.5, "weed": 4.5},
        "individualRecords": {  # W-L-T + points, through Wednesday
            "phil-fowler":   {"team": "mist", "record": "2-1-0", "points": 2},
            "don-fowler":    {"team": "mist", "record": "0-3-0", "points": 0},
            "john-williams": {"team": "mist", "record": "1-2-0", "points": 1},
            "roy-hoenisch":  {"team": "mist", "record": "1-1-1", "points": 1.5},
            "tom-conroy":    {"team": "mist", "record": "1-1-1", "points": 1.5},
            "joe-dueh":      {"team": "mist", "record": "3-0-0", "points": 3},
            "dan-taber":     {"team": "weed", "record": "1-2-0", "points": 1},
            "greg-dueh":     {"team": "weed", "record": "0-3-0", "points": 0},
            "john-taber":    {"team": "weed", "record": "1-2-0", "points": 1},
            "dan-travisano": {"team": "weed", "record": "2-1-0", "points": 2},
            "jim-fowler":    {"team": "weed", "record": "2-0-1", "points": 2.5},
            "craig":         {"team": "weed", "record": "1-1-1", "points": 1.5},
        },
        "result": None,
        "resultNote": ("Final result unknown — the doc only covers through "
                       "Wednesday (4.5-4.5); Thursday singles results were "
                       "not recorded in the folder."),
    }


# ---------------------------------------------------------------- 2009

def extract_2009(alias_lookup):
    """2009: 'HH Score 2009 Working.xls', sheet '2009_Scores'.

    New 18-hole-only layout (37 cols): holes at 5-13/15-23, 18-tot col 25,
    Handicap col 26 (rolling, varies per round), 18-net col 27, plus new
    Pars col 28 / Birdies col 29. Course rating & slope appear in the
    course column on the two rows after each PAR row ('R = 69.7',
    'S = 126'). Real dates at last (Mar 9-13 2009). 8 players; 'Brother
    Bill' = Billy from 2007. Filler rows (blank player, zero scores)
    skipped. Official LeaderBoard sheet exists — places taken from it
    (it breaks net ties: Phil 2nd over Trav, both 382).
    """
    wb = xlrd.open_workbook(HH / "2009" / "HH Score 2009 Working.xls")
    s = wb.sheet_by_name("2009_Scores")
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24))

    rounds = []
    round_nets = {}
    stats = {}

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        assert all(h for h in holes), f"2009 {name}: hole gaps"
        stated_total = to_int(s.cell_value(r, 25))
        assert sum(holes) == stated_total, (
            f"2009 {name}: holes sum {sum(holes)} != stated {stated_total}")

        # rating/slope live in the course column below the PAR row
        cell2 = str(s.cell_value(r, 2))
        m_rating = re.search(r"R\s*=\s*([\d.]+)", cell2)
        m_slope = re.search(r"S\s*=\s*(\d+)", cell2)

        if name == "PAR":
            rounds.append({
                "round": len(rounds) + 1,
                "course": re.sub(r"\s+", " ", s.cell_value(r, 2).strip()),
                "date": xl_date(wb, s.cell_value(r, 4)),
                "holes": 18,
                "par": holes,
                "scores": {},
                "handicaps": {},
            })
        else:
            rd = rounds[-1]
            if m_rating:
                rd["rating"] = float(m_rating.group(1))
            if m_slope:
                rd["slope"] = int(m_slope.group(1))
            pid = alias_lookup[norm_name(name)]
            hcp = to_int(s.cell_value(r, 26))
            net = s.cell_value(r, 27)
            assert abs(stated_total - hcp - net) < 1e-9, (
                f"2009 {name}: net {net} != {stated_total} - {hcp}")
            rd["scores"][pid] = holes
            rd["handicaps"][pid] = hcp
            round_nets[pid] = round_nets.get(pid, 0) + net
            st = stats.setdefault(pid, {"pars": 0, "birdies": 0})
            st["pars"] += to_int(s.cell_value(r, 28)) or 0
            st["birdies"] += to_int(s.cell_value(r, 29)) or 0

    # Official LeaderBoard sheet: Name / Net / Diff / Place.
    lb = wb.sheet_by_name("LeaderBoard")
    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)
    leaderboard = []
    for r in range(1, lb.nrows):
        name = lb.cell_value(r, 0)
        if not str(name).strip():
            continue
        pid = alias_lookup[norm_name(name)]
        net = lb.cell_value(r, 1)
        assert abs(round_nets[pid] - net) < 1e-9, (
            f"2009 LeaderBoard {name}: net {net} != computed {round_nets[pid]}")
        leaderboard.append({"player": pid, "gross": gross[pid], "net": net,
                            "place": to_int(lb.cell_value(r, 3))})

    return {
        "year": 2009,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Dan Taber "
                         "(374, by 8), per the official LeaderBoard sheet."),
        "stats": stats,
    }


# ---------------------------------------------------------------- 2010

def parse_score_sheet_2009_style(wb, sheet_name, year, alias_lookup):
    """Shared parser for the 2009+ 18-hole layout: holes 5-13/15-23,
    18-tot col 25, rolling Handicap col 26, 18-net col 27, Pars/Birdies
    cols 28/29, rating/slope text in the course column, filler rows
    blank. Returns (rounds, round_nets, stats)."""
    s = wb.sheet_by_name(sheet_name)
    HOLE_COLS = list(range(5, 14)) + list(range(15, 24))
    rounds, round_nets, stats = [], {}, {}

    for r in range(1, s.nrows):
        name = s.cell_value(r, 3)
        if not isinstance(name, str) or not name.strip():
            continue
        holes = [to_int(s.cell_value(r, c)) for c in HOLE_COLS]
        if name != "PAR" and all(not h for h in holes):
            continue  # listed player who skipped the trip (all-zero row)
        assert all(h for h in holes), f"{year} {name}: hole gaps"
        stated_total = to_int(s.cell_value(r, 25))
        assert sum(holes) == stated_total, (
            f"{year} {name}: holes sum {sum(holes)} != stated {stated_total}")

        cell2 = str(s.cell_value(r, 2))
        m_rating = re.search(r"R\s*=\s*([\d.]+)", cell2)
        m_slope = re.search(r"S\s*=\s*(\d+)", cell2)

        if name == "PAR":
            rounds.append({
                "round": len(rounds) + 1,
                "course": re.sub(r"\s+", " ", s.cell_value(r, 2).strip()),
                "date": xl_date(wb, s.cell_value(r, 4)),
                "holes": 18,
                "par": holes,
                "scores": {},
                "handicaps": {},
            })
        else:
            rd = rounds[-1]
            if m_rating:
                rd["rating"] = float(m_rating.group(1))
            if m_slope:
                rd["slope"] = int(m_slope.group(1))
            pid = alias_lookup[norm_name(name)]
            hcp = to_int(s.cell_value(r, 26))
            net = s.cell_value(r, 27)
            assert abs(stated_total - hcp - net) < 1e-9, (
                f"{year} {name}: net {net} != {stated_total} - {hcp}")
            rd["scores"][pid] = holes
            rd["handicaps"][pid] = hcp
            round_nets[pid] = round_nets.get(pid, 0) + net
            st = stats.setdefault(pid, {"pars": 0, "birdies": 0})
            st["pars"] += to_int(s.cell_value(r, 28)) or 0
            st["birdies"] += to_int(s.cell_value(r, 29)) or 0
    return rounds, round_nets, stats


def parse_leaderboard_sheet(wb, alias_lookup, round_nets, rounds, year):
    """Official LeaderBoard sheet (Name/Net/Diff/Place) -> leaderboard list,
    with nets asserted against the summed per-round nets."""
    lb = wb.sheet_by_name("LeaderBoard")
    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)
    leaderboard = []
    for r in range(1, lb.nrows):
        name = lb.cell_value(r, 0)
        if not str(name).strip():
            continue
        pid = alias_lookup[norm_name(name)]
        net = lb.cell_value(r, 1)
        assert abs(round_nets[pid] - net) < 1e-9, (
            f"{year} LeaderBoard {name}: net {net} != computed {round_nets[pid]}")
        leaderboard.append({"player": pid, "gross": gross[pid], "net": net,
                            "place": to_int(lb.cell_value(r, 3))})
    return leaderboard


def parse_mist_weed_sheet(wb, alias_lookup, year):
    """Mist-Weed sheet: day blocks of matchups (mist name, _, weed name, _,
    mist pts, weed pts), a running team total, and an individual W-L-T
    table at cols 10-13."""
    s = wb.sheet_by_name("Mist-Weed")
    days, current = [], None
    for r in range(s.nrows):
        c0 = str(s.cell_value(r, 0)).strip()
        if c0.lower().startswith("day"):
            current = {"day": len(days) + 1, "matches": []}
            days.append(current)
            continue
        if current is None or c0 in ("", "Mist"):
            # blank/total/header row; capture day totals when present
            if current is not None and str(s.cell_value(r, 4)).strip() != "":
                v4, v5 = s.cell_value(r, 4), s.cell_value(r, 5)
                if isinstance(v4, float) and c0 == "":
                    current["mistPoints"], current["weedPoints"] = v4, v5
            continue
        weed_name = str(s.cell_value(r, 2)).strip()
        if not weed_name:
            continue
        current["matches"].append({
            "mist": alias_lookup[norm_name(c0)],
            "weed": alias_lookup[norm_name(weed_name)],
            "mistPts": s.cell_value(r, 4),
            "weedPts": s.cell_value(r, 5),
        })

    records = {}
    for r in range(1, s.nrows):
        name = str(s.cell_value(r, 10)).strip()
        if not name:
            continue
        records[alias_lookup[norm_name(name)]] = {
            "wins": to_int(s.cell_value(r, 11)),
            "losses": to_int(s.cell_value(r, 12)),
            "ties": to_int(s.cell_value(r, 13)),
        }

    total_mist = sum(d.get("mistPoints", 0) for d in days)
    total_weed = sum(d.get("weedPoints", 0) for d in days)
    teams = {"mist": sorted({m["mist"] for d in days for m in d["matches"]}),
             "weed": sorted({m["weed"] for d in days for m in d["matches"]})}
    return {
        "teams": teams,
        "days": days,
        "final": {"mist": total_mist, "weed": total_weed},
        "winner": "mist" if total_mist > total_weed else "weed" if total_weed > total_mist else "tie",
        "individualRecords": records,
    }


def extract_2010(alias_lookup):
    """2010: 'HH Score 2010 Working.xls' — 2009-style layout, 12 players,
    Mar 8-12, Myrtle Beach. First in-workbook Mist-Weed sheet (full daily
    match results; final 17.5-12.5 Mist — matches HHScoresHistory).
    Payouts file has the fee structure only (no per-player winnings).
    """
    wb = xlrd.open_workbook(HH / "2010" / "HH Score 2010 Working.xls")
    rounds, round_nets, stats = parse_score_sheet_2009_style(
        wb, "2010 Scores", 2010, alias_lookup)
    leaderboard = parse_leaderboard_sheet(wb, alias_lookup, round_nets, rounds, 2010)
    mist_weed = parse_mist_weed_sheet(wb, alias_lookup, 2010)

    return {
        "year": 2010,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Jim Fowler "
                         "(361, by 10) — his 3rd title. Mist-Weed result "
                         "(Mist 17.5-12.5) also verified."),
        "stats": stats,
        "sideGames": {
            "mistWeed": mist_weed,
            "payoutStructure": ("$18/player: daily low net $12 + daily low team "
                                "$12; champ $2/head; week low net $1/head; "
                                "Mist-Weed $5/player."),
        },
    }


# ---------------------------------------------------------------- 2011

def extract_2011(alias_lookup):
    """2011: 'HH Score 2011 Working.xls' — trip moved to FALL (Oct 16-20).
    8 players; Johnny T / Brother Bill / Knees / Duck are listed with
    all-zero rows (skipped the trip) and are dropped by the shared parser.

    Two stale blocks in this workbook:
      - The LeaderBoard sheet only sorts the template's 8 core slots: it
        ranks the four no-shows on garbage nets and OMITS Gregory & Roy
        (who played). Standings are computed from the score rows instead
        — Gregory's 347 is the low net.
      - The Mist-Weed 'Day Five' block is identical to 2010's day 5
        (absent players, same points) — template leftover. Days 1-4 are
        real (Mist 9.5-6.5). The workbook's running total (13.5-8.5) and
        HHScoresHistory's Team_Hist 2011 row include the stale day; the
        history's own individual records only cover 4 matches/player,
        agreeing with days 1-4. Flagged for Ryan.

    Day 1 match 'Dano v Fill' has its team columns reversed (Fill is
    Mist); corrected here — the fix reconciles the points with both the
    net scores and the individual W-L-T table (Fill 4-0-0).
    """
    wb = xlrd.open_workbook(HH / "2011" / "HH Score 2011 Working.xls")
    rounds, round_nets, stats = parse_score_sheet_2009_style(
        wb, "2011 Scores", 2011, alias_lookup)

    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)
    board = sorted(round_nets.items(), key=lambda kv: kv[1])
    leaderboard = []
    for i, (pid, net) in enumerate(board):
        place = leaderboard[-1]["place"] if leaderboard and leaderboard[-1]["net"] == net else i + 1
        leaderboard.append({"player": pid, "gross": gross[pid], "net": net,
                            "place": place})

    mist_weed = parse_mist_weed_sheet(wb, alias_lookup, 2011)
    # Fix the reversed day-1 match: Fill (Mist) beat Dano (Weed).
    for m in mist_weed["days"][0]["matches"]:
        if m["mist"] == "dan-taber" and m["weed"] == "phil-fowler":
            m["mist"], m["weed"] = m["weed"], m["mist"]
    # Drop the stale 2010 day-5 block and recompute the 4-day final.
    stale_day5 = mist_weed["days"].pop()
    mist_weed["final"] = {
        "mist": sum(d.get("mistPoints", 0) for d in mist_weed["days"]),
        "weed": sum(d.get("weedPoints", 0) for d in mist_weed["days"]),
    }
    mist_weed["winner"] = "mist"
    mist_weed["teams"] = {
        "mist": ["joe-dueh", "tom-conroy", "roy-hoenisch", "phil-fowler"],
        "weed": ["jim-fowler", "greg-dueh", "dan-travisano", "dan-taber"],
    }
    mist_weed["note"] = (
        "Workbook's Day Five block is stale 2010 data (lists players with "
        "all-zero score rows; identical to 2010's day 5) — dropped. Real "
        "days 1-4: Mist 9.5-6.5. NOTE: the workbook running total and "
        "HHScoresHistory Team_Hist both say 13.5-8.5 (22 pts) because they "
        "include the stale day; the history's own individual W-L-T records "
        "(4 matches each) agree with days 1-4 only. Winner is Mist either way.")

    return {
        "year": 2011,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Greg Dueh (347), "
                         "computed low net; the workbook's stale LeaderBoard sheet "
                         "omitted him. Mist-Weed: winner (Mist) is the "
                         "authoritative stat; exact points best-effort."),
        "leaderboardNote": ("Computed from score rows; the workbook LeaderBoard "
                            "sheet only sorted the 8 template slots."),
        "stats": stats,
        "sideGames": {"mistWeed": mist_weed},
    }


# ---------------------------------------------------------------- 2012+ (.xlsm era, openpyxl)

def parse_score_sheet_xlsm(wb, sheet_name, year, alias_lookup):
    """openpyxl port of the 2009-style layout for .xlsm workbooks.
    Same columns; dates are datetimes; blank cells are None."""
    import datetime
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    HOLE_IDX = list(range(5, 14)) + list(range(15, 24))
    rounds, round_nets, stats = [], {}, {}

    for row in rows[1:]:
        name = row[3]
        if not isinstance(name, str) or not name.strip():
            continue
        holes = [to_int(float(row[c])) if row[c] is not None else None for c in HOLE_IDX]
        if name != "PAR" and all(not h for h in holes):
            continue  # no-show row
        assert all(h for h in holes), f"{year} {name}: hole gaps"
        stated_total = to_int(float(row[25]))
        assert sum(holes) == stated_total, (
            f"{year} {name}: holes sum {sum(holes)} != stated {stated_total}")

        cell2 = str(row[2] or "")
        m_rating = re.search(r"R\s*=\s*([\d.]+)", cell2)
        m_slope = re.search(r"S\s*=\s*(\d+)", cell2)

        if name == "PAR":
            rounds.append({
                "round": len(rounds) + 1,
                "course": re.sub(r"\s+", " ", cell2.strip()),
                "date": row[4].strftime("%Y-%m-%d") if isinstance(row[4], datetime.datetime) else str(row[4]),
                "holes": 18,
                "par": holes,
                "scores": {},
                "handicaps": {},
            })
        else:
            rd = rounds[-1]
            if m_rating:
                rd["rating"] = float(m_rating.group(1))
            if m_slope:
                rd["slope"] = int(m_slope.group(1))
            pid = alias_lookup[norm_name(name)]
            hcp = to_int(float(row[26]))
            net = float(row[27])
            assert abs(stated_total - hcp - net) < 1e-9, (
                f"{year} {name}: net {net} != {stated_total} - {hcp}")
            rd["scores"][pid] = holes
            rd["handicaps"][pid] = hcp
            round_nets[pid] = round_nets.get(pid, 0) + net
            st = stats.setdefault(pid, {"pars": 0, "birdies": 0})
            st["pars"] += to_int(float(row[28] or 0)) or 0
            st["birdies"] += to_int(float(row[29] or 0)) or 0
    return rounds, round_nets, stats


def parse_leaderboard_xlsm(wb, alias_lookup, round_nets, rounds, year):
    ws = wb["LeaderBoard"]
    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)
    leaderboard = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        name = row[0]
        if not name or not str(name).strip():
            continue
        pid = alias_lookup[norm_name(str(name))]
        net = float(row[1])
        assert abs(round_nets[pid] - net) < 1e-9, (
            f"{year} LeaderBoard {name}: net {net} != computed {round_nets[pid]}")
        leaderboard.append({"player": pid, "gross": gross[pid], "net": net,
                            "place": to_int(float(row[3]))})
    return leaderboard


def extract_2012(alias_lookup):
    """2012: 'HH Score 2012 Working.xlsm' — first .xlsm year. Fall trip
    Oct 15-19, 12 players. New: Brent, Shelbo; 'Big Gay Loyd' = Roy Hoenisch (identity
    TBD — possibly Roy under a new joke name). Roy/Bruce/Billy/Johnny T
    absent.

    LeaderBoard: TC and Gregory BOTH net 360; sheet places TC 1st.

    Mist-Weed: the workbook sheet is stale 2011 data (includes Roy, who
    didn't attend), and HHScoresHistory's 2012 rows are an exact copy of
    its 2011 rows — no genuine 2012 record survives. (Its Team_Hist 2012
    row, 9.5-6.5/16, is actually 2011's true days-1-4 result, which
    corroborates the 2011 reconstruction.)
    """
    import openpyxl
    wb = openpyxl.load_workbook(
        HH / "2012" / "HH Score 2012 Working.xlsm", read_only=True, data_only=True)
    rounds, round_nets, stats = parse_score_sheet_xlsm(wb, "Scores", 2012, alias_lookup)
    leaderboard = parse_leaderboard_xlsm(wb, alias_lookup, round_nets, rounds, 2012)

    return {
        "year": 2012,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Tom Conroy "
                         "(360, tiebreak over Greg Dueh at 360 per the "
                         "LeaderBoard sheet)."),
        "stats": stats,
        "sideGames": {
            "mistWeed": None,
            "mistWeedNote": ("No genuine 2012 record: the workbook's Mist-Weed "
                             "sheet and HHScoresHistory's 2012 rows are both "
                             "stale copies of 2011 data."),
        },
    }


def parse_mist_weed_xlsm(wb, alias_lookup, year, overrides=None, dm_mode="deadman"):
    """openpyxl port of the Mist-Weed day-block parser.
    dm_mode: how '<carrier> (DM)' rows are credited — 'deadman' (2016) or
    'carrier' (2017, where the individual table gives Deadman 0-0-0)."""
    overrides = overrides or {}

    def pid_of(name):
        key = norm_name(str(name))
        if "(dm" in key:
            if dm_mode == "deadman":   # 2016: DM ball counts as Deadman
                return "deadman"
            key = re.sub(r"\s*\(dm\)?\s*", " ", key).strip()  # 2017: carrier keeps it
        if key in overrides:
            return overrides[key]
        return alias_lookup[key]

    ws = wb["Mist-Weed"]
    rows = list(ws.iter_rows(values_only=True))
    days, current = [], None
    for row in rows:
        c0 = str(row[0] or "").strip()
        if c0.lower().startswith("day"):
            current = {"day": len(days) + 1, "matches": []}
            days.append(current)
            continue
        if current is None or c0 in ("", "Mist"):
            if current is not None and row[4] is not None and c0 == "":
                if isinstance(row[4], (int, float)):
                    current["mistPoints"], current["weedPoints"] = float(row[4]), float(row[5])
            continue
        weed_name = str(row[2] or "").strip()
        if not weed_name:
            continue
        m_pts, w_pts = float(row[4]), float(row[5])
        if m_pts == 0 and w_pts == 0:
            continue  # unplayed template row (e.g. 2014 day 5)
        current["matches"].append({
            "mist": pid_of(c0),
            "weed": pid_of(weed_name),
            "mistPts": m_pts,
            "weedPts": w_pts,
        })

    records = {}
    for row in rows[1:]:
        name = str(row[10] or "").strip()
        if not name or row[11] is None or not isinstance(row[11], (int, float)):
            continue
        records[pid_of(name)] = {
            "wins": to_int(float(row[11])),
            "losses": to_int(float(row[12])),
            "ties": to_int(float(row[13])),
        }

    days = [d for d in days if d["matches"]]  # drop unplayed template days
    for i, d in enumerate(days):
        d["day"] = i + 1
    total_mist = sum(d.get("mistPoints", 0) for d in days)
    total_weed = sum(d.get("weedPoints", 0) for d in days)
    return {
        "teams": {"mist": sorted({m["mist"] for d in days for m in d["matches"]}),
                  "weed": sorted({m["weed"] for d in days for m in d["matches"]})},
        "days": days,
        "final": {"mist": total_mist, "weed": total_weed},
        "winner": "mist" if total_mist > total_weed else "weed" if total_weed > total_mist else "tie",
        "individualRecords": records,
    }


def extract_2013(alias_lookup):
    """2013: 'HH Score 2013 Working.xlsm' — standardized era. Sep 23-27,
    12 players (Joey absent, Bruce back). LeaderBoard: Jimmy and Brent
    BOTH 361; sheet places Jimmy 1st.
    Mist-Weed complete (Mist 16-14, matches Team_Hist). The Mist-Weed
    sheet calls Big Gay Loyd 'Roy' — Ryan confirmed 2026-07-08 that
    'Big Gay Loyd' WAS Roy Hoenisch all along (there is no Loyd).
    """
    import openpyxl
    wb = openpyxl.load_workbook(
        HH / "2013" / "HH Score 2013 Working.xlsm", read_only=True, data_only=True)
    rounds, round_nets, stats = parse_score_sheet_xlsm(wb, "Scores", 2013, alias_lookup)
    leaderboard = parse_leaderboard_xlsm(wb, alias_lookup, round_nets, rounds, 2013)
    mist_weed = parse_mist_weed_xlsm(wb, alias_lookup, 2013,
                                     overrides={"roy": "roy-hoenisch"})
    mist_weed["note"] = ("'Roy' / 'Big Gay Loyd' = Roy Hoenisch (Ryan "
                         "confirmed 2026-07-08: there is no Loyd). "
                         "Also: the per-match grid gives "
                         "Brent 4-1/Phil 1-3-1 but the official W-L-T table "
                         "(and HHScoresHistory) says Brent 3-2/Phil 2-2-1 — a "
                         "pairing row was misrecorded; the table is authoritative, "
                         "team points unaffected.")

    return {
        "year": 2013,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Jim Fowler "
                         "(361, tiebreak over Brent at 361). 4th title."),
        "stats": stats,
        "sideGames": {"mistWeed": mist_weed},
    }


MW_OVERRIDES_2014 = {
    "roy": "roy-hoenisch",
    "hitman": "t-dog-hitman", "t-dog": "t-dog-hitman",
    "cookie": "cookie-ac-milan", "ac milan": "cookie-ac-milan",
}


def extract_2014(alias_lookup):
    """2014: 'HH Score 2014 Working.xlsm'. Oct 20-24, 12 players.
    New: 'Cookie AC Milan' and 'T-Dog Hitman' (real names TBD);
    Duck/Knees/Bruce out. Shelbo wins two years after finishing last.
    Mist-Weed: Weed's first-ever win, 18-6 (Mist swept 0-6 on day 1;
    day 5 not played/recorded — history's 24-pt total agrees). The
    sheet again calls him 'Roy' (= Roy Hoenisch)."""
    import openpyxl
    wb = openpyxl.load_workbook(
        HH / "2014" / "HH Score 2014 Working.xlsm", read_only=True, data_only=True)
    rounds, round_nets, stats = parse_score_sheet_xlsm(wb, "Scores", 2014, alias_lookup)
    leaderboard = parse_leaderboard_xlsm(wb, alias_lookup, round_nets, rounds, 2014)
    mist_weed = parse_mist_weed_xlsm(wb, alias_lookup, 2014,
                                     overrides=MW_OVERRIDES_2014)

    return {
        "year": 2014,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Dan Shelbo Sr. "
                         "(355, by 3) — last place 2012, champion 2014."),
        "stats": stats,
        "sideGames": {"mistWeed": mist_weed},
    }


def parse_lostball_xlsm(wb, alias_lookup, overrides=None):
    """LostBall sheet: two columns, player / number, total row at bottom."""
    overrides = overrides or {}
    ws = wb["LostBall"]
    out = {}
    for row in ws.iter_rows(values_only=True):
        name = str(row[0] or "").strip()
        if not name or row[1] is None or not isinstance(row[1], (int, float)):
            continue
        key = norm_name(name)
        if key in ("player", "no player", "total", "deadman", "actual"):
            continue  # header / placeholder / actual-count rows
        pid = overrides.get(key) or alias_lookup[key]
        out[pid] = to_int(float(row[1]))
    return out


def extract_2015(alias_lookup):
    """2015: 'HH Score 2015 Working.xlsm'. Sep 20-24, 12 players (Cookie
    out, Knees back). Champion T-Dog Hitman 337 by 18 — dead last in
    2014, champion 2015. First LostBall sheet (numbers captured verbatim;
    interpretation unclear — history's Shots sheet says 125 lost balls
    for 2015 but the sheet totals 1828 — Ryan to explain).
    Mist-Weed: Mist 20.5-9.5 (matches Team_Hist). Sheet calls Roy 'BGR'
    / 'Roy'. The individual W-L-T table only covers days 1-4 (24
    matches); HHScoresHistory copied the same 4-day table."""
    import openpyxl
    wb = openpyxl.load_workbook(
        HH / "2015" / "HH Score 2015 Working.xlsm", read_only=True, data_only=True)
    overrides = {"roy": "roy-hoenisch", "bgr": "roy-hoenisch",
                 "t-dog": "t-dog-hitman", "hitman": "t-dog-hitman"}
    rounds, round_nets, stats = parse_score_sheet_xlsm(wb, "Scores", 2015, alias_lookup)
    leaderboard = parse_leaderboard_xlsm(wb, alias_lookup, round_nets, rounds, 2015)
    mist_weed = parse_mist_weed_xlsm(wb, alias_lookup, 2015, overrides=overrides)
    mist_weed["note"] = ("Individual W-L-T table (and HHScoresHistory's copy of "
                         "it) only covers days 1-4; the match grid has all 5 "
                         "days. Team result includes day 5. Sheet shorthand "
                         "'BGR'/'Roy' = Roy Hoenisch.")
    lost_ball = parse_lostball_xlsm(wb, alias_lookup, overrides=overrides)

    return {
        "year": 2015,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = T-Dog Hitman "
                         "(337, by 18) — last in 2014, champion 2015. LostBall "
                         "numbers parked per Ryan (stored verbatim, unlabeled)."),
        "stats": stats,
        "sideGames": {
            "mistWeed": mist_weed,
            "lostBall": lost_ball,
            "lostBallNote": ("First year of the LostBall sheet. Numbers stored "
                             "verbatim — unit unclear (history says 125 balls "
                             "lost in 2015; sheet totals 1828). Ryan to explain."),
        },
    }


def extract_2016(alias_lookup):
    """2016: 'HH Score 2016 working.xlsm'. Sep 26-30, Brunswick-area
    courses. 12 'players' — but 'Deadman' looks like a phantom entry:
    Mist-Weed matches recorded as '<carrier> (DM)' with a rotating Mist
    player carrying the ball, 0 lost balls, yet full scores and 1st on
    the LeaderBoard (368, tiebreak over Phil 368). New real players:
    Jeff, Troy. Mist-Weed: Weed 16.5-13.5 (matches Team_Hist); the
    W-L-T table covers all 5 days this year (new Sum column)."""
    import openpyxl
    wb = openpyxl.load_workbook(
        HH / "2016" / "HH Score 2016 working.xlsm", read_only=True, data_only=True)
    overrides = {"roy": "roy-hoenisch", "bgr": "roy-hoenisch", "jef": "jeff",
                 "t-dog": "t-dog-hitman", "hitman": "t-dog-hitman"}
    rounds, round_nets, stats = parse_score_sheet_xlsm(wb, "Scores", 2016, alias_lookup)

    # LeaderBoard verbatim, but Deadman has NO recorded scores (all-zero
    # rows) — his 368 is unverifiable; every real player's net is asserted.
    ws = wb["LeaderBoard"]
    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)
    # 'Deadman' (sheet place 1, net 368) was a placeholder for the missing
    # 12th player — Ryan ruled 2026-07-04 the trip was 11 players and the
    # real champion is Phil. Dropped from the leaderboard; places shift up.
    leaderboard = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        name = row[0]
        if not name or not str(name).strip():
            continue
        pid = alias_lookup[norm_name(str(name))]
        if pid == "deadman":
            continue
        net = float(row[1])
        assert abs(round_nets[pid] - net) < 1e-9, (
            f"2016 LeaderBoard {name}: net {net} != computed {round_nets[pid]}")
        leaderboard.append({"player": pid, "gross": gross[pid], "net": net,
                            "place": len(leaderboard) + 1})

    mist_weed = parse_mist_weed_xlsm(wb, alias_lookup, 2016, overrides=overrides)
    mist_weed["note"] = ("Deadman's matches are recorded as '<carrier> (DM)' — "
                         "a rotating Mist player carried the Deadman ball each "
                         "day; all such rows are credited to 'deadman'. "
                         "HHScoresHistory instead folded some DM results into "
                         "the carriers' records (TC 5-1, Roy 3-3, Shelbo 1-5 "
                         "there) and has no Deadman row — workbook table kept "
                         "here; team result identical either way.")
    lost_ball = parse_lostball_xlsm(wb, alias_lookup, overrides=overrides)

    return {
        "year": 2016,
        "location": "Myrtle Beach / Brunswick, NC-SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Phil Fowler "
                         "(368). The sheet's 'Deadman' (also 368, sheet place 1) "
                         "was a placeholder for the missing 12th player — 11 "
                         "real players in 2016; dropped from the leaderboard."),
        "stats": stats,
        "sideGames": {"mistWeed": mist_weed, "lostBall": lost_ball},
    }


def extract_2017(alias_lookup):
    """2017: 'HH Score 2017 Working.xlsm' (daily snapshots also exist;
    Working == Working_Fri). Sep 18-22: Arrowhead, Tidewater, Barefoot
    Fazio, Thistle, Prestwick. 11 real players + the Deadman placeholder
    again (all-zero scores, sheet place 1 at 'net 0' — dropped; real
    champion Trav 367 on the tiebreak over Roy 367).
    Mist-Weed: only days 1-4 played (day 5 all zeros — dropped);
    final Weed 12.5-11.5 (matches Team_Hist 24-pt year). This year the
    '(DM)' rows count for the CARRIER (individual table: Deadman 0-0-0,
    Roy 6 matches) — dm_mode='carrier'. Row 'T-Dog(D2)Troy' = Troy's
    match (tie), per the table's Sum column."""
    import openpyxl
    wb = openpyxl.load_workbook(
        HH / "2017" / "HH Score 2017 Working.xlsm", read_only=True, data_only=True)
    overrides = {"roy": "roy-hoenisch", "bgr": "roy-hoenisch", "jef": "jeff",
                 "t-dog": "t-dog-hitman", "hitman": "t-dog-hitman",
                 "t-dog(d2)troy": "troy"}
    rounds, round_nets, stats = parse_score_sheet_xlsm(wb, "Scores", 2017, alias_lookup)

    ws = wb["LeaderBoard"]
    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)
    leaderboard = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        name = row[0]
        if not name or not str(name).strip():
            continue
        pid = alias_lookup[norm_name(str(name))]
        if pid == "deadman":  # placeholder row (net 0)
            continue
        net = float(row[1])
        assert abs(round_nets[pid] - net) < 1e-9, (
            f"2017 LeaderBoard {name}: net {net} != computed {round_nets[pid]}")
        leaderboard.append({"player": pid, "gross": gross[pid], "net": net,
                            "place": len(leaderboard) + 1})

    mist_weed = parse_mist_weed_xlsm(wb, alias_lookup, 2017,
                                     overrides=overrides, dm_mode="carrier")
    mist_weed["days"] = [d for d in mist_weed["days"]
                         if any(m["mistPts"] or m["weedPts"] for m in d["matches"])]
    mist_weed["final"] = {
        "mist": sum(d.get("mistPoints", 0) for d in mist_weed["days"]),
        "weed": sum(d.get("weedPoints", 0) for d in mist_weed["days"]),
    }
    mist_weed["winner"] = "weed"
    mist_weed["individualRecords"] = {
        pid: rec for pid, rec in mist_weed["individualRecords"].items()
        if pid != "deadman"}
    mist_weed["note"] = ("Only days 1-4 were played (day-5 block all zeros — "
                         "dropped). '(DM)' rows credited to the carrier this "
                         "year per the workbook's own W-L-T table (Deadman "
                         "0-0-0 there).")
    lost_ball = parse_lostball_xlsm(wb, alias_lookup, overrides=overrides)

    return {
        "year": 2017,
        "location": "Myrtle Beach, SC",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Dan Travisano "
                         "(367, tiebreak over Roy Hoenisch 367). 2nd title."),
        "stats": stats,
        "sideGames": {"mistWeed": mist_weed, "lostBall": lost_ball},
    }


def extract_2018(alias_lookup):
    """2018: 'HH Score 2018 Working.xlsm' ('Woring' = day-1 snapshot).
    Trip moves NORTH: Ocean City MD / Delaware courses, Oct 15-19.
    12 players: Duck & Brent back, Shelbo/T-Dog/Joe out; NEW 'Jimmy B'
    (Jim Fowler becomes 'Jimmy F'). Champion Trav 351 — back-to-back.
    Mist-Weed: 4 days, Mist 13-11 (matches Team_Hist). '(DM)' rows are
    carrier-credited (table: Deadman 0-0-0); 'Deadman (TC)' = TC.
    Jimmy B did not play Mist-Weed."""
    import openpyxl
    wb = openpyxl.load_workbook(
        HH / "2018" / "HH Score 2018 Working.xlsm", read_only=True, data_only=True)
    overrides = {"roy": "roy-hoenisch", "bgr": "roy-hoenisch", "jef": "jeff",
                 "t-dog": "t-dog-hitman", "hitman": "t-dog-hitman",
                 "deadman (tc)": "tom-conroy", "jimmt f": "jim-fowler"}
    rounds, round_nets, stats = parse_score_sheet_xlsm(wb, "Scores", 2018, alias_lookup)
    leaderboard = parse_leaderboard_xlsm(wb, alias_lookup, round_nets, rounds, 2018)
    mist_weed = parse_mist_weed_xlsm(wb, alias_lookup, 2018,
                                     overrides=overrides, dm_mode="carrier")
    mist_weed["individualRecords"].pop("deadman", None)
    mist_weed["note"] = ("4-day competition. '(DM)' rows carrier-credited per "
                         "the workbook table; 'Deadman (TC)' row = TC. "
                         "Jimmy B sat out Mist-Weed.")
    lost_ball = parse_lostball_xlsm(wb, alias_lookup, overrides=overrides)

    return {
        "year": 2018,
        "location": "Ocean City, MD area (MD/DE)",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Dan Travisano "
                         "(351, by 7). Back-to-back, 3rd title."),
        "stats": stats,
        "sideGames": {"mistWeed": mist_weed, "lostBall": lost_ball},
    }


def extract_2019(alias_lookup):
    """2019: 'HH Score 2019 Working.xlsm'. Ocean City MD again, Sep 16-19
    2019 (dates embedded as 'Date:' strings in the course column; the
    Date-column serials are 2018 template leftovers). Round 5 (9/21,
    course 'None') never played — all zeros, dropped. 8 players; 'No
    Player' placeholder rows dropped. Champion Duck 283 over 4 rounds
    (1st title). Mist-Weed: 3 days only, Mist 9-3; stale day-4 template
    block dropped. HHScoresHistory's tracking ends at 2018, so the
    workbook is the sole Mist-Weed source from here on."""
    import openpyxl
    wb = openpyxl.load_workbook(
        HH / "2019" / "HH Score 2019 Working.xlsm", read_only=True, data_only=True)
    overrides = {"roy": "roy-hoenisch", "bgr": "roy-hoenisch",
                 "t-dog": "t-dog-hitman", "hitman": "t-dog-hitman"}
    rounds, round_nets, stats = parse_score_sheet_xlsm(wb, "Scores", 2019, alias_lookup)

    rounds = [rd for rd in rounds if rd["scores"]]  # drop unplayed round 5
    real_dates = ["2019-09-16", "2019-09-17", "2019-09-18", "2019-09-19"]
    for rd, d in zip(rounds, real_dates):
        rd["date"] = d

    ws = wb["LeaderBoard"]
    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)
    leaderboard = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        name = str(row[0] or "").strip()
        if not name or norm_name(name) == "no player":
            continue
        pid = alias_lookup[norm_name(name)]
        net = float(row[1])
        assert abs(round_nets[pid] - net) < 1e-9, (
            f"2019 LeaderBoard {name}: net {net} != computed {round_nets[pid]}")
        leaderboard.append({"player": pid, "gross": gross[pid], "net": net,
                            "place": len(leaderboard) + 1})

    mist_weed = parse_mist_weed_xlsm(wb, alias_lookup, 2019,
                                     overrides=overrides, dm_mode="carrier")
    mist_weed["days"] = [d for d in mist_weed["days"]
                         if any(m["mistPts"] or m["weedPts"] for m in d["matches"])]
    mist_weed["final"] = {
        "mist": sum(d.get("mistPoints", 0) for d in mist_weed["days"]),
        "weed": sum(d.get("weedPoints", 0) for d in mist_weed["days"]),
    }
    mist_weed["winner"] = "mist"
    mist_weed["teams"] = {
        "mist": ["tom-conroy", "don-fowler", "troy", "phil-fowler"],
        "weed": ["greg-dueh", "jim-fowler", "dan-taber", "dan-travisano"],
    }
    mist_weed["note"] = ("3-day competition (day-4 block was stale 2018 "
                         "template data — dropped). Not in HHScoresHistory "
                         "(its tracking ends 2018).")
    lost_ball = parse_lostball_xlsm(wb, alias_lookup, overrides=overrides)

    return {
        "year": 2019,
        "location": "Ocean City, MD area",
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "championNote": ("Verified by Ryan 2026-07-04: champion = Don Fowler "
                         "(283 over 4 rounds, by 7). First title."),
        "stats": stats,
        "sideGames": {"mistWeed": mist_weed, "lostBall": lost_ball},
    }


def _extract_oc_era(year, path, real_dates, date_note, alias_lookup,
                    extra_overrides=None, location="Ocean City, MD area",
                    tolerate_lb_mismatch=False):
    """Shared extractor for the 2019+ Ocean City-era workbooks: drop
    unplayed rounds ('None' course, zero rows), drop 'No Player'
    placeholder leaderboard rows, Mist-Weed in carrier mode, LostBall
    verbatim."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    overrides = {"roy": "roy-hoenisch", "bgr": "roy-hoenisch",
                 "t-dog": "t-dog-hitman", "hitman": "t-dog-hitman"}
    overrides.update(extra_overrides or {})
    rounds, round_nets, stats = parse_score_sheet_xlsm(wb, "Scores", year, alias_lookup)

    rounds = [rd for rd in rounds if rd["scores"]]
    if real_dates:
        for rd, d in zip(rounds, real_dates):
            rd["date"] = d

    ws = wb["LeaderBoard"]
    gross = {}
    for rd in rounds:
        for pid, holes in rd["scores"].items():
            gross[pid] = gross.get(pid, 0) + sum(holes)
    leaderboard = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        name = str(row[0] or "").strip()
        if not name or norm_name(name) in ("no player", "deadman"):
            continue
        pid = overrides.get(norm_name(name)) or alias_lookup[norm_name(name)]
        if pid not in round_nets:  # listed but no scores recorded (e.g. Scotty H 2025)
            continue
        net = float(row[1])
        entry = {"player": pid, "gross": gross[pid], "net": net,
                 "place": len(leaderboard) + 1}
        if abs(round_nets[pid] - net) > 1e-9:
            if not tolerate_lb_mismatch:
                raise AssertionError(
                    f"{year} LeaderBoard {name}: net {net} != computed {round_nets[pid]}")
            entry["computedNet"] = round_nets[pid]
        leaderboard.append(entry)

    mist_weed = parse_mist_weed_xlsm(wb, alias_lookup, year,
                                     overrides=overrides, dm_mode="carrier")
    mist_weed["days"] = [d for d in mist_weed["days"]
                         if any(m["mistPts"] or m["weedPts"] for m in d["matches"])]
    mist_weed["final"] = {
        "mist": sum(d.get("mistPoints", 0) for d in mist_weed["days"]),
        "weed": sum(d.get("weedPoints", 0) for d in mist_weed["days"]),
    }
    m, w = mist_weed["final"]["mist"], mist_weed["final"]["weed"]
    mist_weed["winner"] = "mist" if m > w else "weed" if w > m else "tie"
    mist_weed["individualRecords"].pop("deadman", None)
    lost_ball = parse_lostball_xlsm(wb, alias_lookup, overrides=overrides)

    return {
        "year": year,
        "location": location,
        "courses": [rd["course"] for rd in rounds],
        "scoring": "net",
        "handicapSystem": "Rolling daily handicaps (per-round, in rounds[].handicaps).",
        "rounds": rounds,
        "leaderboard": leaderboard,
        "champion": leaderboard[0]["player"],
        "stats": stats,
        "dateNote": date_note,
        "sideGames": {"mistWeed": mist_weed, "lostBall": lost_ball},
    }


def extract_2020(alias_lookup):
    """2020 (COVID year — trip still ran): 'HH Score 2020 Working.xlsm'
    is the final (Working_2/3/3A are daily snapshots). Ocean City, Sep
    21-24, 4 rounds (R5 'None' dropped). 11 players + 'No Player'
    placeholder. 'BGR' mapped to Roy Hoenisch (recap-style naming) —
    pending Ryan. Deadman ghost on Weed with named carriers
    ('Deadman (Trav)' etc.), carrier-credited. Champion Shelbo 262 by 22.
    Mist-Weed: Weed 12.5-11.5."""
    ex = _extract_oc_era(
        2020, HH / "2020" / "HH Score 2020 Working.xlsm",
        ["2020-09-21", "2020-09-22", "2020-09-23", "2020-09-24"],
        "Dates from the score sheet (Sep 21-24 2020, consistent with file timestamps).",
        alias_lookup,
        extra_overrides={"bgr": "roy-hoenisch", "roy": "roy-hoenisch",
                         "fhil": "phil-fowler", "jamie": "jim-fowler",
                         "deadman (trav)": "dan-travisano",
                         "deadman (dano)": "dan-taber",
                         "deadman (jimmy)": "jim-fowler"},
        tolerate_lb_mismatch=True)
    ex["sideGames"]["lostBallNote"] = (
        "2020 LostBall sheet is a 'Lost Guess' pool (each player's guess "
        "at the week's total lost balls); actual per-day lost balls are "
        "in the Totals sheet LB block — week total 111. 'Jamie' = Jimmy F.")
    ex["championNote"] = ("Verified by Ryan 2026-07-04: champion = Dan Shelbo "
                          "Sr. (262, by 22 — 2nd title). BGR = Roy Hoenisch "
                          "(Ryan confirmed).")
    ex["leaderboardNote"] = (
        "Official LeaderBoard day-4 nets differ from the Scores sheet's "
        "net column (day-4 handicaps diverge; Scores file was recalculated "
        "in 2022). LBHistory days 1-3 match Scores exactly. The official "
        "LeaderBoard totals are kept as 'net'; the Scores-sheet sum is in "
        "'computedNet' where different. Champion unaffected (Shelbo wins "
        "by 22 official / 16 computed).")
    ex["sideGames"]["mistWeed"]["note"] = (
        "4 days. Deadman ghost played for Weed with named carriers "
        "('Deadman (Trav/Dano/Jimmy)') — carrier-credited per the "
        "workbook's W-L-T table.")
    return ex


def extract_2021(alias_lookup):
    """2021: 'HH Score 2021 Working_Final.xlsm'. Ocean City, 4 rounds +
    an unscored scramble day (recap doc). 12 players (Kevin debuts;
    Junior = Dan Shelbo Jr.). Champion Matt 289 — CONFIRMED by the
    'HHGA_Recap_2021' doc ('Overall Winners: Jacket - Matt'), which also
    confirms Mist-Weed Weed 13-11 and 135 lost balls. Score-sheet dates
    are 2020 template serials; real dates ~Sep 20-23 2021 (estimated)."""
    ex = _extract_oc_era(
        2021, HH / "2021" / "HH Score 2021 Working_Final.xlsm",
        ["2021-09-20", "2021-09-21", "2021-09-22", "2021-09-23"],
        "ESTIMATED Sep 20-23 2021 (sheet kept 2020 template dates; file saved Fri Sep 24 2021).",
        alias_lookup,
        extra_overrides={"bgr": "roy-hoenisch", "roy": "roy-hoenisch",
                         "shelbo jr": "dan-shelbo-jr", "jamie": "jim-fowler"},
        tolerate_lb_mismatch=True)
    ex["leaderboardNote"] = (
        "Official LeaderBoard kept as 'net'; small divergences vs the "
        "Scores sheet's net column (file recalculated Sep 2022) recorded "
        "as 'computedNet'. Champion confirmed by the recap doc regardless.")
    ex["championNote"] = ("Verified by Ryan 2026-07-04 (and HHGA_Recap_2021: "
                          "'Jacket - Matt'): champion = Matt (289, by 4). "
                          "Monkey's Ass = Kevin. BGR = Roy Hoenisch (Ryan "
                          "confirmed).")
    ex["sideGames"]["mistWeed"]["note"] = (
        "4 days, Weed 13-11 — matches the recap doc. A scramble day was "
        "also played (not individually scored, not in Scores).")
    return ex


def extract_2022(alias_lookup):
    """2022: 'HH Score 2022 Working.xlsm'. Ocean City, Sep 19-22, 4
    rounds (R5 'None' dropped). 12 players — Joe Dueh RETURNS after 8
    years. Champion Shelbo Jr. 299 on the tiebreak over Duck 299 —
    confirmed by the published Leader_Board_2022_04.pdf in the folder.
    Mist-Weed: 3 days only (day-4 block zeros, dropped), Weed 12.5-5.5.
    'Danp' typo = Dano."""
    ex = _extract_oc_era(
        2022, HH / "2022" / "HH Score 2022 Working.xlsm",
        ["2022-09-19", "2022-09-20", "2022-09-21", "2022-09-22"],
        "Dates from the score sheet (Sep 19-22 2022).",
        alias_lookup,
        extra_overrides={"bgr": "roy-hoenisch", "roy": "roy-hoenisch",
                         "shelbo jr": "dan-shelbo-jr", "jamie": "jim-fowler",
                         "danp": "dan-taber"},
        tolerate_lb_mismatch=True)
    ex["championNote"] = ("Verified by Ryan 2026-07-04 (and the published "
                          "Leader_Board_2022_04.pdf): champion = Dan Shelbo Jr. "
                          "(299, tiebreak over Duck 299). First title.")
    ex["sideGames"]["mistWeed"]["note"] = (
        "3 days played (day-4 block all zeros — dropped). Weed 12.5-5.5.")
    return ex


def extract_2023(alias_lookup):
    """2023: 'HH Score 2023 Working.xlsm' (daily snapshots 09_14..09_17
    also in folder → real dates Sep 14-17 2023; the sheet kept 2022
    template serials). 4 rounds (R5 'None' dropped): Heritage Shores,
    River Run, Man O War, Ocean Pines. 12 players (Kevin back, Matt out).
    Champion Shelbo Sr. 293 by 1 over Duck. Mist-Weed: 4 days, Weed
    14.5-9.5 — Weed's 4th straight."""
    ex = _extract_oc_era(
        2023, HH / "2023" / "HH Score 2023 Working.xlsm",
        ["2023-09-14", "2023-09-15", "2023-09-16", "2023-09-17"],
        "Dates from the daily snapshot filenames (09_14..09_17); sheet kept 2022 template serials.",
        alias_lookup,
        extra_overrides={"bgr": "roy-hoenisch", "roy": "roy-hoenisch",
                         "shelbo jr": "dan-shelbo-jr", "jamie": "jim-fowler"},
        tolerate_lb_mismatch=True)
    ex["championNote"] = ("Verified by Ryan 2026-07-04: champion = Dan Shelbo "
                          "Sr. (293, by 1 over Duck). 3rd title; Shelbo family "
                          "back-to-back.")
    ex["sideGames"]["mistWeed"]["note"] = "4 days. Weed 14.5-9.5 — 4th straight Weed win."
    return ex


def extract_2024(alias_lookup):
    """2024: the real data lives in 'HH Score 2024_Day 4 Working.xlsm' —
    the plain 'Working.xlsm' is an unfilled template shell (empty scores,
    2023's Mist-Weed copied in, stale LeaderBoard). Ocean City, Sep
    18-21: Glen Riddle Man of War, Glen Riddle War Admiral, Bear Trap
    Dunes, Baywood Greens. 12 players — RYAN (Fowler) debuts.
    Champion Shelbo Sr. 276 by 7 ('new trial format' per the Year in
    Review, which confirms the whole leaderboard + Mist-Weed Weed
    13.5-10.5 + Kevin Burke's full name)."""
    ex = _extract_oc_era(
        2024, HH / "2024" / "HH Score 2024_Day 4 Working.xlsm",
        ["2024-09-18", "2024-09-19", "2024-09-20", "2024-09-21"],
        "Dates from the score sheet (Sep 18-21 2024).",
        alias_lookup,
        extra_overrides={"bgr": "roy-hoenisch", "roy": "roy-hoenisch",
                         "shelbo jr": "dan-shelbo-jr", "jamie": "jim-fowler"},
        tolerate_lb_mismatch=True)
    ex["championNote"] = ("Verified by Ryan 2026-07-04 (and the 2024 Year in "
                          "Review): champion = Dan Shelbo Sr. 276, by 7 over "
                          "Shelbo Jr. — 4th title. The 'new trial format' = "
                          "winner by cumulative net over the 4 days, PGA-style "
                          "(per Ryan; prior format unclear).")
    ex["sideGames"]["mistWeed"]["note"] = (
        "4 days, Weed 13.5-10.5 (5th straight Weed win) — matches the Year "
        "in Review, including all individual W-L-T records.")
    ex["sourceNote"] = ("Extracted from the Day 4 working file; the main "
                        "'HH Score 2024 Working.xlsm' is an unfilled shell.")
    return ex


def extract_2025(alias_lookup):
    """2025: 'HH Score 2025_Working.xlsm'. TRIP MOVES TO ATLANTIC CITY,
    NJ (Brigantine house): Vineyard National (Renault), Harbor Pines,
    Seaview Bay, Twisted Dune — Oct 1-4 2025 (real serials). 11 scored
    players + 'Scotty H' listed with no scores (dropped, pending Ryan);
    the Year in Review also says Greg's and Jimmy's scores weren't
    recorded. Champion BGR/Roy Hoenisch 288 — 'the drought is broken'
    (YIR-confirmed; first title since joining 2006). Deadman ghost back
    in DEADMAN-credited mode ('Deadman (BGR/Troy)' rows; YIR lists
    Deadman 2-2-0). Mist-Weed: Weed 13-11, 6th straight (YIR ✓)."""
    ex = _extract_oc_era(
        2025, HH / "2025" / "HH Score 2025_Working.xlsm",
        ["2025-10-01", "2025-10-02", "2025-10-03", "2025-10-04"],
        "Dates from the score sheet (Oct 1-4 2025).",
        alias_lookup,
        extra_overrides={"bgr": "roy-hoenisch", "roy": "roy-hoenisch",
                         "shelbo jr": "dan-shelbo-jr", "jamie": "jim-fowler",
                         "scotty h": "scotty-h",
                         "deadman (bgr)": "deadman",
                         "deadman (troy)": "deadman",
                         "deaman (troy)": "deadman"},
        location="Atlantic City / Brigantine, NJ",
        tolerate_lb_mismatch=True)
    ex["championNote"] = ("Verified by Ryan 2026-07-04 (and the 2025 Year in "
                          "Review): champion = Roy Hoenisch 288, by 5 — 'The "
                          "drought is broken'. Monkey's Ass: Dano.")
    # the shared helper drops Deadman, but the 2025 YIR credits the ghost
    ex["sideGames"]["mistWeed"]["individualRecords"]["deadman"] = {
        "wins": 2, "losses": 2, "ties": 0}
    ex["sideGames"]["mistWeed"]["note"] = (
        "4 days, Weed 13-11 (6th straight) — matches the Year in Review. "
        "Deadman ghost played for Mist with carriers ('Deadman (BGR/Troy)'), "
        "credited to Deadman per the YIR table (2-2-0). NOTE: the workbook/YIR "
        "table says Troy 2-2-0 but his own four matches read 1-3-0 — kept as "
        "recorded.")
    ex["rosterNote"] = ("'Scotty H' listed on the roster with no recorded "
                        "scores (dropped from leaderboard). YIR: 'We did not "
                        "record Greg or Jimmy scores.' Ryan to clarify who "
                        "attended.")
    return ex


# ------------------------------------------------- Jimmy's archive (2026-07)
# Source: Jimmy_Data/'HHGA Web site design.doc' (champions list = FACT per
# Ryan) and 'Stats and calculations.xlsx' (Champ Round / Final Four flags).
# The 1999-2009 champion was decided at a final CHAMPIONSHIP COURSE among a
# Final Four — not by cumulative net. 2000-02 were gross-score years counted
# on the 18-hole rounds (proven: Jimmy & Dano tie at 471 in 2000 -> the only
# playoff in HHGA history).

CHAMPION_CORRECTIONS = {
    2001: ("jim-fowler", "Jimmy — 'wears down the field and wins with a 90 at "
           "Shipyard... three years out of 4, Jimmy makes it three for three.'"),
    2002: ("phil-fowler", "Phil — rookie year: 'pummels the field and shoots a 90 "
           "at Shipyard', triggering the handicap era."),
    2004: ("tom-conroy", "TC — 'hits two 30 footers down the stretch to pull "
           "away from Phil.'"),
    2006: ("joe-dueh", "Joe — 'wins his first of two at King's North applying "
           "the 2 year plan to perfection.'"),
    2009: ("phil-fowler", "Phil — 'Long Bay-Gate': picks the championship course, "
           "then wins it there."),
}

YEAR_HISTORY = {  # championship-format block + Jimmy's stories, 1999-2009
    2000: {"finalFour": ["dan-taber", "jim-fowler", "john-taber", "joe-dueh"],
           "championshipCourse": "Shipyard", "format": "gross (18-hole rounds)",
           "story": ("The only playoff in HHGA history: Jimmy and Dano tie at 471 "
                     "gross, and Jimmy wins with a bogey on Shipyard #10. 'It was "
                     "tense, but I was up for the challenge,' recalls Jimmy.")},
    2001: {"finalFour": ["jim-fowler", "tom-conroy", "dan-taber", "joe-dueh"],
           "championshipCourse": "Shipyard", "format": "gross (18-hole rounds)",
           "story": ("Jimmy wears down the field and wins with a 90 at Shipyard — "
                     "three titles in a row. Johnny gets called home from Crescent "
                     "Point, sick child.")},
    2002: {"finalFour": ["phil-fowler", "joe-dueh", "jim-fowler", "tom-conroy"],
           "championshipCourse": "Shipyard", "format": "gross (18-hole rounds)",
           "story": ("The days of innocence are over: rookie Phil pummels the field "
                     "and shoots a 90 at Shipyard — the last year gross scores crown "
                     "a champion. Phil brings a wretched pink jacket as the trophy. "
                     "TC's first swing puts his new 3-wood clubhead in the pond; fear "
                     "of alligators keeps him from retrieving it. Jose retires.")},
    2003: {"finalFour": ["jim-fowler", "dan-taber", "john-williams", "greg-dueh"],
           "championshipCourse": "Arthur Hill",
           "story": ("Pappy leverages the first year of handicap scoring to crush the "
                     "field — low net every single day, a beatdown never replicated. "
                     "Joe skips for Hawaii with his wife. Last year at Hilton Head; "
                     "Pappy sells the condo. We love Pappy.")},
    2004: {"finalFour": ["phil-fowler", "tom-conroy", "jim-fowler", "dan-taber"],
           "championshipCourse": "MBN West", "warmups": ["MBN King's North"],
           "story": ("TC hits two 30-footers down the stretch to pull away from Phil. "
                     "After Pappy sells the Hilton Head condo, Trav opens up Myrtle "
                     "Beach and a whole new world for the HHGA.")},
    2005: {"finalFour": ["dan-taber", "greg-dueh", "dan-travisano", "jim-fowler"],
           "championshipCourse": "Wachesaw East", "warmups": ["Wizard"],
           "story": ("In the heat and wind at Wachesaw East, Greg's patented low "
                     "boring shot takes the jacket. First year in two condos; the "
                     "Florida contingent arrives: Bruce Morris and his friend Al, "
                     "who quickly earns the nickname 'Big Gay' and never returns.")},
    2006: {"finalFour": ["joe-dueh", "tom-conroy", "dan-travisano", "john-taber"],
           "championshipCourse": "Kings North", "warmups": ["Man O War", "Wizard"],
           "story": ("Joe wins his first of two at King's North, applying the "
                     "'2-year plan' to perfection. Big Gay Al didn't fit in and "
                     "didn't return — his replacement, the one and only Roy "
                     "Hoenisch: indirectly related by marriage, the much-heralded "
                     "4th Fowler on the tour.")},
    2007: {"finalFour": ["john-taber", "bruce-morris", "dan-taber", "john-williams"],
           "championshipCourse": "Mn O War", "warmups": ["World Turd", "River Oaks"],
           "story": ("An early downpour washes out Pappy and the competition fades "
                     "on familiar Man O War — Johnny birdies #18 in style with the "
                     "gallery watching to win his first title. Floridian Craig "
                     "'Knees' joins the tour. Owl on Witch #2; Billy's grunting "
                     "boar imitation.")},
    2008: {"finalFour": ["phil-fowler", "joe-dueh", "dan-taber", "craig"],
           "championshipCourse": "MBN North", "warmups": ["MBN West"],
           "story": ("King's North becomes Joe's favorite course as he wins his 2nd "
                     "championship at the tricked-up tract — 'like a fish in water.' "
                     "Subsequent meetings result in King's North being forever "
                     "banned from championship-course play.")},
    2009: {"finalFour": ["phil-fowler", "dan-taber", "jim-fowler", "john-taber"],
           "championshipCourse": "Long Bay",
           "warmups": ["MBN Southcreek", "MBN West"],
           "story": ("'Long Bay-Gate': Phil picks the championship course and then "
                     "proceeds to win it there. Warmest week ever, drinks by the "
                     "pool, down to 8 diehards — the economy taking its toll.")},
}

EARLY_YEARS = [
    {
        "year": 1998,
        "location": "Fayetteville, NC",
        "courses": [],
        "rounds": [],
        "leaderboard": [],
        "attendees": ["jim-fowler", "dan-taber", "tom-conroy", "john-taber",
                      "john-williams", "jose-perna", "joe-dueh"],
        "champion": None,
        "championNote": "No champion — total washout (Jimmy's history, fact per Ryan).",
        "story": ("The year it all started (for Jimmy anyway). Fayetteville, North "
                  "Carolina: a total washout — it rained every day. Forever known "
                  "as 'Fayette-nam'. But it started a tradition of golf "
                  "man-cations."),
        "noScores": True,
    },
    {
        "year": 1999,
        "location": "Hilton Head Island, SC",
        "courses": ["Old Carolina", "Old South", "Shipyard"],
        "coursesNote": "Partial — two courses unknown ('Old Carolina, Old South, ??, ??, Shipyard').",
        "rounds": [],
        "leaderboard": [],
        "attendees": ["jim-fowler", "dan-taber", "tom-conroy", "john-taber",
                      "john-williams", "jose-perna", "joe-dueh"],
        "champion": "jim-fowler",
        "championNote": ("From Jimmy's history (fact per Ryan): first-ever HHGA "
                         "championship. No scores survive."),
        "championship": {
            "finalThree": ["jim-fowler", "john-taber", "joe-dueh"],
            "championshipCourse": "Shipyard",
        },
        "story": ("Jimmy outlasts Joe and Johnny at Shipyard — the round is lost "
                  "for Joe when a single is put in the final threesome, Johnny "
                  "fades early, and Jimmy cruises to the first ever HHGA "
                  "championship. First year as guests at Pappy's condo: the condo "
                  "is cool, the hot tub is warm. Van from Johnny's to Philly, fly "
                  "to Savannah. The start of it all — the rest is history."),
        "noScores": True,
    },
]


def apply_jimmy_history(years):
    """Champion corrections + championship blocks from Jimmy's archive,
    plus the 1998/1999 story-only years."""
    for yr in years:
        y = yr["year"]
        if y in CHAMPION_CORRECTIONS:
            pid, note = CHAMPION_CORRECTIONS[y]
            yr["champion"] = pid
            yr["championNote"] = (
                f"Champion per Jimmy's 'HHGA Web site design.doc' (fact, Ryan "
                f"2026-07-08): {note} Supersedes the earlier cumulative-net "
                f"reading. Leaderboard below = cumulative standings, not the "
                f"championship result.")
        if y in YEAR_HISTORY:
            h = YEAR_HISTORY[y]
            yr["championship"] = {
                "finalFour": h["finalFour"],
                "championshipCourse": h["championshipCourse"],
            }
            if "format" in h:
                yr["championship"]["format"] = h["format"]
            if "warmups" in h:
                yr["warmups"] = h["warmups"]
            yr["story"] = h["story"]
    return EARLY_YEARS + years


# ---------------------------------------------------------------- main

EXTRACTORS = {
    2000: extract_2000,
    2001: extract_2001,
    2002: extract_2002,
    2003: extract_2003,
    2004: extract_2004,
    2005: extract_2005,
    2006: extract_2006,
    2007: extract_2007,
    2008: extract_2008,
    2009: extract_2009,
    2010: extract_2010,
    2011: extract_2011,
    2012: extract_2012,
    2013: extract_2013,
    2014: extract_2014,
    2015: extract_2015,
    2016: extract_2016,
    2017: extract_2017,
    2018: extract_2018,
    2019: extract_2019,
    2020: extract_2020,
    2021: extract_2021,
    2022: extract_2022,
    2023: extract_2023,
    2024: extract_2024,
    2025: extract_2025,
}


def main():
    players, alias_lookup = load_aliases()
    YEARS_DIR.mkdir(parents=True, exist_ok=True)

    # One clean file per year (for review / diffing), then the combined
    # hhga.json the website actually fetches.
    years = []
    for y in sorted(EXTRACTORS):
        yr = EXTRACTORS[y](alias_lookup)
        years.append(yr)

    years = apply_jimmy_history(years)
    for yr in years:
        year_file = YEARS_DIR / f"{yr['year']}.json"
        year_file.write_text(json.dumps(yr, indent=2))
        print(f"Wrote {year_file} ({year_file.stat().st_size:,} bytes) — "
              f"{len(yr['rounds'])} rounds, {len(yr['leaderboard'])} players, "
              f"champion={yr['champion']}")

    data = {"players": players, "years": years}
    OUT.write_text(json.dumps(data, indent=2))
    print(f"Wrote {OUT} ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
